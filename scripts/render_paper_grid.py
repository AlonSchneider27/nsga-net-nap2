"""Render paper-grid xlsx/csv + linear-axis plots from normalized_report.json.

Usage: python scripts/render_paper_grid.py <dir containing normalized_report.json>
(run scripts/normalize_native_runs.py first). Needs openpyxl + matplotlib."""
import json, os, csv
import matplotlib; matplotlib.use('Agg'); import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
OUT = sys.argv[1] if len(sys.argv) > 1 else '.'
R = json.load(open(f'{OUT}/normalized_report.json'))
LC = ['sotl', 'sotl_e', 'early_stop', 'lce_m', 'lc_pfn']; ZC = ['synflow', 'grad_norm', 'snip']
PAPER = [1, 2, 3, 7, 11, 17, 23]; NAPB = [1, 2, 3, 5, 7, 11, 17, 23]
ORDER = [t for t in ['in_c10', 'in_c100', 'in_in16', 'x_c100_to_c10', 'x_in16_to_c10', 'x_c10_to_c100', 'x_in16_to_c100', 'x_c10_to_in16', 'x_c100_to_in16'] if t in R]
COL = {'nap2': '#d62728', 'sotl_e': '#1f77b4', 'sotl': '#17becf', 'lc_pfn': '#2ca02c', 'lce_m': '#9467bd',
       'early_stop': '#ff7f0e', 'synflow': '#8c8c8c', 'grad_norm': '#bcbcbc', 'snip': '#d9d9d9'}

HF = PatternFill('solid', fgColor='305496'); HFONT = Font(color='FFFFFF', bold=True, size=10)
BOLD = Font(bold=True, color='1F4E00'); ITAL = Font(italic=True, color='7F7F7F'); NA = Font(color='BBBBBB')
THIN = Side(style='thin', color='D9D9D9'); B = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SEC = PatternFill('solid', fgColor='E2EFDA')

def write_grid(ws, row0, title, grid, note):
    ws.cell(row0, 1, title).font = Font(bold=True, size=11); row0 += 1
    ws.cell(row0, 1, note).font = Font(italic=True, size=9, color='888888'); row0 += 1
    hdr = ['method'] + [f'Q={k}' for k in PAPER]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row0, j, h); c.fill = HF; c.font = HFONT; c.alignment = Alignment(horizontal='center'); c.border = B
    row0 += 1
    colmax = {k: max((g[f'@{k}'] for g in grid if g[f'@{k}'] is not None), default=None) for k in PAPER}
    for g in grid:
        ws.cell(row0, 1, g['method']).border = B
        for j, k in enumerate(PAPER, 2):
            v, fl = g[f'@{k}'], g[f'@{k}_flag']
            c = ws.cell(row0, j, v if v is not None else 'n/a'); c.border = B; c.alignment = Alignment(horizontal='center')
            if v is None: c.font = NA
            elif colmax[k] is not None and abs(v - colmax[k]) < 1e-9: c.font = BOLD
            elif fl == 'interp': c.font = ITAL
        row0 += 1
    return row0 + 1

wb = Workbook(); wb.remove(wb.active)
csv_rows = []
for tag in ORDER:
    r = R[tag]; ws = wb.create_sheet(r['label'].replace('->', 'to'))
    ws.column_dimensions['A'].width = 12
    for col in 'BCDEFGHIJKLMN': ws.column_dimensions[col].width = 9
    row = 1
    ws.cell(row, 1, f"{r['label']}  —  n={r['n']} unique archs, seed 42").font = Font(bold=True, size=12); row += 1
    ws.cell(row, 1, (f"Timing (this run's GPU): epoch = {r['T_epoch']} s train ({r['T_val']} s val pass); "
                     f"nap2 snapshot = {r['T_snap']} s (23-snapshot query {r['T_nap2_23']} s incl. stats/AE/predict). "
                     f"Seconds-normalization: 1 epoch = {r['r']} nap2-snapshot-equivalents.")).font = Font(italic=True, size=9); row += 2
    note = ("Rows = nap2 snapshot budgets (paper Tables 4-6 query times). LC methods placed on the seconds axis (epoch K -> K x r); "
            "italic = linearly interpolated between adjacent native epochs; n/a = budget below the first epoch. Bold = column best.")
    row = write_grid(ws, row, 'KT vs 20-epoch valid_acc (this run\'s proxy GT)', r['grid_valid'], note)
    for g in r['grid_valid']:
        csv_rows.append({'run': r['label'], 'gt': 'valid_acc_20ep', 'method': g['method'], **{f'Q={k}': g[f'@{k}'] for k in PAPER}, **{f'Q={k}_flag': g[f'@{k}_flag'] for k in PAPER}})
    if r['grid_nb201']:
        row = write_grid(ws, row, 'KT vs NB-201 200-epoch test accuracy (the paper\'s GT)', r['grid_nb201'], note)
        for g in r['grid_nb201']:
            csv_rows.append({'run': r['label'], 'gt': 'nb201_200ep', 'method': g['method'], **{f'Q={k}': g[f'@{k}'] for k in PAPER}, **{f'Q={k}_flag': g[f'@{k}_flag'] for k in PAPER}})
    # native points block
    ws.cell(row, 1, 'Native observation points (what was actually measured)').font = Font(bold=True, size=11); row += 1
    hdr = ['method', 'epoch K', 'snap-eq (K x r)', 'seconds', 'KT (valid_acc)'] + (['KT (NB201)'] if r['grid_nb201'] else [])
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row, j, h); c.fill = SEC; c.font = Font(bold=True, size=9); c.border = B
    row += 1
    for m in LC:
        nb = dict(r['nb201_native'][m]) if r['nb201_native'] else {}
        for K, x, s, kt in r['native_valid'][m]:
            vals = [m, K, x, s, kt] + ([nb.get(K)] if r['grid_nb201'] else [])
            for j, v in enumerate(vals, 1):
                c = ws.cell(row, j, v if v is not None else 'n/a'); c.border = B
            row += 1
    for k, s, kt in r['nap2_native']:
        vals = ['nap2', f'snap {k}', k, s, kt] + ([r['nb201_nap2'].get(str(k)) if r['nb201_nap2'] else None] if r['grid_nb201'] else [])
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v if v is not None else 'n/a'); c.border = B
        row += 1
    for m in ZC:
        vals = [m, 'init', 0, 0, r['zc_valid'][m]] + ([r['nb201_zc'][m]] if r['grid_nb201'] else [])
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v if v is not None else 'n/a'); c.border = B
        row += 1
    ws.freeze_panes = 'A4'
wb.save(f'{OUT}/kt_paper_grid.xlsx')
with open(f'{OUT}/kt_paper_grid.csv', 'w', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(csv_rows[0].keys())); w.writeheader(); w.writerows(csv_rows)

# ---- plots: linear x = nap2-snapshot-equivalents (evenly spaced), LC points at K*r ----
def panel(ax, tag, gtkey):
    r = R[tag]; vsrc = r['native_valid'] if gtkey == 'valid' else r['nb201_native']
    for m in LC:
        pts = [(K * r['r'], kt) for K, x, s, kt in r['native_valid'][m]] if gtkey == 'valid' else \
              [(K * r['r'], kt) for K, kt in r['nb201_native'][m]]
        pts = [(x, y) for x, y in pts if y is not None and x <= 23.5]
        if pts: ax.plot(*zip(*pts), marker='o', ms=3.5, lw=1.2, color=COL[m], label=m)
    nap = [(k, kt) for k, s, kt in r['nap2_native']] if gtkey == 'valid' else [(k, r['nb201_nap2'][str(k)]) for k in NAPB]
    nap = [(x, y) for x, y in nap if y is not None]
    ax.plot(*zip(*nap), marker='s', ms=3.5, lw=2.0, color=COL['nap2'], label='nap2', zorder=5)
    zc = r['zc_valid'] if gtkey == 'valid' else r['nb201_zc']
    for m in ZC:
        if zc.get(m) is not None: ax.axhline(zc[m], ls=':', lw=0.9, color=COL[m], label=m)
    ax.set_xlim(0, 24); ax.set_xticks(PAPER); ax.set_ylim(0.1, 0.8); ax.grid(alpha=0.25); ax.tick_params(labelsize=7)
    ax.set_title(f"{r['label']} (n={r['n']}; 1 epoch = {r['r']} snap)", fontsize=9)
    ax.set_xlabel('query budget in nap2-snapshot equivalents (seconds-normalized)', fontsize=7)

import math
NC = 4; NR = math.ceil((len(ORDER) + 1) / NC)
fig, axes = plt.subplots(NR, NC, figsize=(17, 3.75 * NR)); axes = axes.ravel()
for ax, tag in zip(axes, ORDER): panel(ax, tag, 'valid')
for ax in axes[len(ORDER):]: ax.axis('off')
h, l = axes[0].get_legend_handles_labels(); axes[len(ORDER)].legend(h, l, loc='center', fontsize=8, title='methods')
axes[0].set_ylabel('KT vs 20-epoch valid_acc', fontsize=8)
fig.suptitle('Seconds-normalized (Michael): LC epochs mapped to nap2-snapshot equivalents; linear axis with paper budgets as ticks', fontsize=11)
fig.tight_layout(); fig.savefig(f'{OUT}/kt_normalized_valid_acc.png', dpi=140); plt.close(fig)

fig, axes = plt.subplots(NR, NC, figsize=(17, 3.75 * NR)); axes = axes.ravel()
for ax, tag in zip(axes, ORDER): panel(ax, tag, 'nb201')
for ax in axes[len(ORDER):]: ax.axis('off')
h, l = axes[0].get_legend_handles_labels(); axes[len(ORDER)].legend(h, l, loc='center', fontsize=8, title='methods')
axes[0].set_ylabel('KT vs NB-201 200-epoch test acc', fontsize=8)
fig.suptitle("Seconds-normalized, vs the paper's ground truth (NB-201 200-epoch accuracy) — all 7 runs", fontsize=11)
fig.tight_layout(); fig.savefig(f'{OUT}/kt_normalized_nb201_gt.png', dpi=140); plt.close(fig)
print('wrote', os.listdir(OUT))
